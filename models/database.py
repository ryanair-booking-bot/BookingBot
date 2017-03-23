"""This module does blah blah."""
# pylint: disable=too-few-public-methods

import os
import datetime
import pandas as pd
from openpyxl import load_workbook

class Database(object):
    "Shared connection to the database"

    def __init__(self):
        self.cities = ["Warsaw", "Manchester", "Berlin"]

        os.chdir('./models')
        self.workbook = load_workbook('./FlightListIE2016.xlsx')
        first_sheet = self.workbook.get_sheet_names()[0]
        self.flights = self.workbook.get_sheet_by_name(first_sheet)

    def does_place_exist(self, destination_name):
        "Checks if place exists"
        if destination_name in self.cities:
            return True
        return False

    def get_flights(self, departure_station, arrival_station, departure_date):
        "ss"

        result = []
        for row in range(2, self.flights.max_row):
            source = self.flights.cell(row=row, column=3).value
            destination = self.flights.cell(row=row, column=4).value
            date = self.flights.cell(row=row, column=1).value
            if source == departure_station and destination == arrival_station \
                and date == departure_date:
                result.append([x.value for x in self.flights[row]])
        return result
